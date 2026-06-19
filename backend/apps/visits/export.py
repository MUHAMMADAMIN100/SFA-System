from django.http import HttpResponse
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from apps.users.permissions import IsAdmin

from .models import Visit
from .views import filter_visits_by_date

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

HEADERS = [
    "Дата и время",
    "Менеджер",
    "Магазин",
    "Товар",
    "Объём/вес",
    "Отгружено",
    "Просрочка",
]


@extend_schema(
    summary="Экспорт визитов в Excel (xlsx)",
    parameters=[
        OpenApiParameter("date_from", OpenApiTypes.DATE, OpenApiParameter.QUERY),
        OpenApiParameter("date_to", OpenApiTypes.DATE, OpenApiParameter.QUERY),
    ],
    responses={(200, XLSX_CONTENT_TYPE): OpenApiTypes.BINARY},
)
class VisitsExportView(APIView):
    """Экспорт визитов за период в xlsx (одна строка на товар)."""

    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        queryset = (
            filter_visits_by_date(Visit.objects.all(), request.query_params)
            .select_related("manager", "store")
            .prefetch_related("items__product")
            .order_by("created_at")
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Визиты"

        bold = Font(bold=True)
        for col, title in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=title)
            cell.font = bold

        row_idx = 2
        for visit in queryset:
            manager_name = visit.manager.full_name or visit.manager.username
            created = visit.created_at.strftime("%d.%m.%Y %H:%M")
            for item in visit.items.all():
                sheet.cell(row=row_idx, column=1, value=created)
                sheet.cell(row=row_idx, column=2, value=manager_name)
                sheet.cell(row=row_idx, column=3, value=visit.store.name)
                sheet.cell(row=row_idx, column=4, value=item.product.name)
                sheet.cell(row=row_idx, column=5, value=item.product.volume)
                sheet.cell(row=row_idx, column=6, value=item.shipped_qty)
                sheet.cell(row=row_idx, column=7, value=item.expired_qty)
                row_idx += 1

        widths = [18, 24, 28, 28, 12, 12, 12]
        for col, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = width

        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response["Content-Disposition"] = 'attachment; filename="visits.xlsx"'
        workbook.save(response)
        return response
