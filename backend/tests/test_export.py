import io

import pytest
from openpyxl import load_workbook

pytestmark = pytest.mark.django_db

EXPORT_URL = "/api/export/visits/"
XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_returns_xlsx(admin_client, manager_user, stores, products, make_visit):
    make_visit(manager_user, stores[0], [(products[0], 10, 2), (products[1], 5, 0)])

    resp = admin_client.get(EXPORT_URL)
    assert resp.status_code == 200
    assert resp["Content-Type"] == XLSX_CT
    assert "attachment" in resp["Content-Disposition"]
    assert resp["Content-Disposition"].endswith('.xlsx"')


def test_export_content_parses_and_has_rows(
    admin_client, manager_user, stores, products, make_visit
):
    make_visit(manager_user, stores[0], [(products[0], 10, 2), (products[1], 5, 0)])

    resp = admin_client.get(EXPORT_URL)
    wb = load_workbook(io.BytesIO(resp.content))
    sheet = wb.active

    header = [c.value for c in sheet[1]]
    assert header == [
        "Дата и время",
        "Менеджер",
        "Магазин",
        "Товар",
        "Объём/вес",
        "Отгружено",
        "Просрочка",
    ]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2  # одна строка на товар

    by_product = {r[3]: r for r in rows}
    milk = by_product["Молоко"]
    assert milk[1] == "Тест Менеджер"
    assert milk[2] == "Магазин А"
    assert milk[4] == "1 л"
    assert milk[5] == 10  # отгружено
    assert milk[6] == 2   # просрочка


def test_export_respects_date_filter(
    admin_client, manager_user, stores, products, make_visit
):
    from datetime import timedelta

    from django.utils import timezone

    now = timezone.now()
    old = now - timedelta(days=400)
    make_visit(manager_user, stores[0], [(products[0], 99, 9)], created_at=old)
    make_visit(manager_user, stores[0], [(products[0], 1, 0)], created_at=now)

    resp = admin_client.get(EXPORT_URL, {"date_from": now.date().isoformat()})
    wb = load_workbook(io.BytesIO(resp.content))
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][5] == 1
